import pandas as pd
import matplotlib.pyplot as plt
import os
import numpy as np
from configparser import ConfigParser
from openpyxl import load_workbook
import openpyxl
from datetime import datetime

def ganntplot_config(configini):
    dateTime = datetime.now().strftime('%Y%m%d%H%M%S')
    config = ConfigParser()
    configFile = str(configini)
    config.read(configFile)
    frames_dir_out = config.get('Frame settings', 'frames_dir_out')
    frames_dir_out = os.path.join(frames_dir_out, 'gantt_plots')
    if not os.path.exists(frames_dir_out):
        os.makedirs(frames_dir_out)
    csv_dir = config.get('General settings', 'csv_path')
    csv_dir_in = os.path.join(csv_dir, 'machine_results')
    no_targets = config.getint('SML settings', 'No_targets')
    use_master = config.get('General settings', 'use_master_config')
    vidInfPath = config.get('General settings', 'project_path')
    vidInfPath = os.path.join(vidInfPath, 'logs')
    vidInfPath = os.path.join(vidInfPath, 'video_info.csv')
    vidinfDf = pd.read_csv(vidInfPath)
    boutEnd = 0
    boutStart = 0
    boutEnd_list = [0]
    boutStart_list = []
    filesFound = []
    target_names = []
    data_event = []
    log_list = []
    colours = ['red', 'green', 'pink', 'orange', 'blue', 'purple', 'lavender', 'grey', 'sienna', 'tomato', 'azure',
               'crimson', 'aqua', 'plum', 'teal', 'maroon', 'lime', 'coral']
    colourTupleX = list(np.arange(3.5, 203.5, 5))
    configFilelist = []

    loop = 0
    loopy = 0

    ########### FIND CSV FILES ###########
    if use_master == 'yes':
        for i in os.listdir(csv_dir_in):
            if i.__contains__(".csv"):
                file = os.path.join(csv_dir_in, i)
                filesFound.append(file)
    if use_master == 'no':
        config_folder_path = config.get('General settings', 'config_folder')
        for i in os.listdir(config_folder_path):
            if i.__contains__(".ini"):
                configFilelist.append(os.path.join(config_folder_path, i))
                iniVidName = i.split(".")[0]
                csv_fn = iniVidName + '.csv'
                file = os.path.join(csv_dir_in, csv_fn)
                filesFound.append(file)

    ########### GET TARGET COLUMN NAMES ###########
    for ff in range(no_targets):
        currentModelNames = 'target_name_' + str(ff + 1)
        currentModelNames = config.get('SML settings', currentModelNames)
        currentModelNames = currentModelNames  + '_prediction'
        target_names.append(currentModelNames)
    colours = colours[:len(target_names)]

    ########### logfile path ###########
    log_fn = config.get('General settings', 'project_name')
    log_fn = log_fn + '.xlsx'
    log_path = config.get('General settings', 'project_path')
    log_path = os.path.join(log_path, 'project_folder', 'logs')
    log_fn = os.path.join(log_path, log_fn)
    if not os.path.exists(log_path):
        os.makedirs(log_path)
    if not os.path.isfile(log_fn):
        wb = openpyxl.Workbook()
        wb.save(log_fn)

    sheetname = 'sklearn_' + dateTime
    headers = ['Video']
    for i in target_names:
        head1 = str(i) + '_event_Nos'
        head2 = str(i) + '_Tot_event_duration'
        head3 = str(i) + '_mean_event_duration'
        head4 = str(i) + '_median_event_duration'
        head5 = str(i) + '_time_to_first_occurance'
        head6 = str(i) + '_mean_event_interval'
        head7 = str(i) + '_median_event_interval'
        headers.extend([head1, head2, head3, head4, head5, head6, head7])
    log_df = pd.DataFrame(columns=headers)

    for i in filesFound:
        boutsDf = pd.DataFrame(columns=['Event', 'Start_frame', 'End_frame'])
        currentFile = i
        if use_master == 'no':
            configFile = configFilelist[loopy]
            config = ConfigParser()
            config.read(configFile)
            fps = config.getint('Frame settings', 'fps')
        CurrentVideoName = os.path.basename(currentFile)
        videoSettings = vidinfDf.loc[vidinfDf['Video'] == str(CurrentVideoName.replace('.csv', ''))]
        fps = int(videoSettings['fps'])
        loopy += 1
        readDf = pd.read_csv(currentFile)
        currCol = [col for col in readDf.columns if 'prediction' in col]
        currCol = [x for x in currCol if "Probability" not in x]
        dataDf = readDf.filter(currCol, axis=1)
        dataDf['frames'] = readDf['frames']
        rowCount = dataDf.shape[0]
        folderNm = os.path.basename(currentFile)
        logFolderNm = 'Video' + str(folderNm.split('.')[0])
        folderName = str(folderNm.split('.')[0]) + str('_gantt')
        saveDir = os.path.join(frames_dir_out, folderName)
        if not os.path.exists(saveDir):
            os.makedirs(saveDir)
        for bb in currCol:
            currTarget = bb
            for indexes, rows in dataDf[dataDf['frames'] >= boutEnd].iterrows():
                if rows[currTarget] == 1:
                    boutStart = rows['frames']
                    for index, row in dataDf[dataDf['frames'] >= boutStart].iterrows():
                        if row[currTarget] == 0:
                            boutEnd = row['frames']
                            if boutEnd_list[-1] != boutEnd:
                                boutStart_list.append(boutStart)
                                boutEnd_list.append(boutEnd)
                                values = [currTarget, boutStart, boutEnd]
                                boutsDf.loc[(len(boutsDf))] = values
                                break
                            break
            boutStart_list = [0]
            boutEnd_list = [0]
            boutEnd = 0
            boutStart = 0

        # Convert to time
        boutsDf['Start_time'] = boutsDf['Start_frame'] / fps
        boutsDf['End_time'] = boutsDf['End_frame'] / fps
        boutsDf['Bout_time'] = boutsDf['End_time'] - boutsDf['Start_time']

        # record logs
        log_list = []
        log_list.append(logFolderNm)
        for i in target_names:
            currDf = boutsDf.loc[boutsDf['Event'] == i]
            try:
                firstOccur = currDf['Start_time'].iloc[0]
            except IndexError:
                firstOccur = 0
            eventNOs = len(currDf)
            TotEventDur = round(currDf['Bout_time'].sum(), 2)
            try:
                MeanEventDur = TotEventDur / eventNOs
                MedianEventDur = currDf['Bout_time'].median()
            except ZeroDivisionError:
                MeanEventDur = 0
                MedianEventDur = 0
            currDf_shifted = currDf.shift(periods=-1)
            currDf_shifted = currDf_shifted.drop(columns=['Event', 'Start_frame', 'End_frame', 'End_time', 'Bout_time'])
            currDf_shifted = currDf_shifted.rename(columns={'Start_time': 'Start_time_shifted'})
            currDf_combined = pd.concat([currDf, currDf_shifted], axis=1, join='inner')
            currDf_combined['Event_interval'] = currDf_combined['Start_time_shifted'] - currDf_combined['End_time']
            meanEventInterval = currDf_combined["Event_interval"].mean()
            medianEventInterval = currDf_combined['Bout_time'].median()
            log_list.append(eventNOs)
            log_list.append(TotEventDur)
            log_list.append(MeanEventDur)
            log_list.append(MedianEventDur)
            log_list.append(firstOccur)
            log_list.append(meanEventInterval)
            log_list.append(medianEventInterval)
        log_df.loc[loop] = log_list
        print(log_df)

        ################### PLOT #######################
        loop = 0
        relRows = pd.DataFrame()
        for k in range(rowCount):
            fig, ax = plt.subplots()
            ylabels = ([s.replace('_prediction', '') for s in target_names])
            # ylabels = np.array([ylabels])
            currentDf = dataDf.iloc[:k]
            relRows = boutsDf.loc[boutsDf['End_frame'] <= k]
            for i, event in enumerate(relRows.groupby("Event")):
                for x in target_names:
                    if event[0] == x:
                        ix = target_names.index(x)
                        data_event = event[1][["Start_time", "Bout_time"]]
                        ax.broken_barh(data_event.values, (colourTupleX[ix], 3), facecolors=colours[ix])
                        loop += 1
            xLength = (round(k / fps)) + 1
            if xLength < 10:
                xLength = 10
            loop = 0
            ax.set_xlim(0, xLength)
            ax.set_ylim(0, colourTupleX[len(target_names)])
            setYtick = np.arange(5, 5 * no_targets + 1, 5)
            ax.set_yticks(setYtick)
            ax.set_yticklabels(ylabels, rotation=45, fontsize=8)
            ax.yaxis.grid(True)
            filename = (str(k) + '.png')
            savePath = os.path.join(saveDir, filename)
            plt.savefig(savePath)
            print('Saved gantt plot ' + str(k))
            plt.close('all')
        loop += 1

    book = load_workbook(log_fn)
    writer = pd.ExcelWriter(log_fn, engine='openpyxl')
    writer.book = book
    log_df = log_df.fillna(0)
    log_df.to_excel(writer, sheet_name=sheetname, index=False)
    writer.save()
    writer.close()
    print('Finihed generating gantt plot.')