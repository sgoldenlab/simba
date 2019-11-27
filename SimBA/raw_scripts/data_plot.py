import os
import pandas as pd
import re
import statistics
import numpy as np
import cv2
from configparser import ConfigParser
from openpyxl import load_workbook
import openpyxl
from datetime import datetime
dateTime = datetime.now().strftime('%Y%m%d%H%M%S')

config = ConfigParser()
configFile = r"Z:\DeepLabCut\DLC_extract\New_082119\project_folder\project_config.ini"
config.read(configFile)

frames_dir_out = config.get('Frame settings', 'frames_dir_out')
frames_dir_out = os.path.join(frames_dir_out, 'live_data_table')
if not os.path.exists(frames_dir_out):
    os.makedirs(frames_dir_out)
csv_dir = config.get('General settings', 'csv_path')
csv_dir_in = os.path.join(csv_dir, 'machine_results')
use_master = config.get('General settings', 'use_master_config')
vidInfPath = config.get('General settings', 'project_path')
vidInfPath = os.path.join(vidInfPath, 'project_folder', 'logs')
vidInfPath = os.path.join(vidInfPath, 'video_info.csv')
vidinfDf = pd.read_csv(vidInfPath)
filesFound = []
loop = 0
loopy = 0
VideoNo_list = []
OneMMpixel = config.getint('Frame settings', 'mm_per_pixel')

########### logfile path ###########
log_fn = config.get('General settings', 'project_name')
log_fn = log_fn + '.xlsx'
log_path = config.get('General settings', 'project_path')
log_path = os.path.join(log_path, 'project_folder', 'logs')
log_fn = os.path.join(log_path, log_fn)
if not os.path.exists(log_path):
    os.makedirs(log_path)
if not os.path.isfile(log_fn):
    wb = openpyxl.Workbook()
    wb.save(log_fn)

sheetname = 'Data_log_' + dateTime
columns = ['Video', 'Mean_velocity_CD1', 'Mean_velocity_C57', 'Total_movement_CD1', 'Total_movement_C57', 'Mean_centroid_distance', 'Mean_nose_to_nose_distance']
log_df = pd.DataFrame(columns=columns)
configFilelist = []

########### FIND CSV FILES ###########
if use_master == 'yes':
    for i in os.listdir(csv_dir_in):
        if i.__contains__(".csv"):
            file = os.path.join(csv_dir_in, i)
            filesFound.append(file)
if use_master == 'no':
    config_folder_path = config.get('General settings', 'config_folder')
    for i in os.listdir(config_folder_path):
        if i.__contains__(".ini"):
            configFilelist.append(os.path.join(config_folder_path, i))
            iniVidName = i.split(".")[0]
            csv_fn = iniVidName + '.csv'
            file = os.path.join(csv_dir_in, csv_fn)
            filesFound.append(file)

frames_processed_list = []
meanVeloM1 = []
meanVeloM2 = []
totMoveM1 = []
totMoveM2 = []
mean_centroid_distance_mm = []
mean_nose_distance_mm = []


for i in filesFound:
    centroid_distance_mm_list = []
    nose_2_nose_dist_mm_list = []
    frameCounter = 0
    meanVelocity_M1_list = []
    meanVelocity_M2_list = []
    total_Movement_M1_list = []
    total_Movement_M2_list = []
    list_nose_movement_M1 = []
    list_nose_movement_M2 = []
    loop = 0
    currentFile = i
    if use_master == 'no':
        configFile = configFilelist[loopy]
        config = ConfigParser()
        config.read(configFile)
        fps = config.getint('Frame settings', 'fps')
        OneMMpixel = config.getint('Frame settings', 'mm_per_pixel')
    CurrentVideoName = os.path.basename(currentFile)
    videoSettings = vidinfDf.loc[vidinfDf['Video'] == str(CurrentVideoName.replace('.csv', ''))]
    fps = int(videoSettings['fps'])
    OneMMpixel = int(videoSettings['pixels/mm'])
    loopy+=1
    csv_df = pd.read_csv(currentFile)
    VideoNo = os.path.basename(currentFile)
    VideoNo = 'Video' + str(re.sub("[^0-9]", "", VideoNo))
    VideoNo_list.append(VideoNo)
    imagesDirOut = VideoNo + str('_data_plots')
    savePath = os.path.join(frames_dir_out, imagesDirOut)
    if not os.path.exists(savePath):
        os.makedirs(savePath)
    df_lists = [csv_df[i:i+fps] for i in range(0,csv_df.shape[0],fps)]


    for i in df_lists:
        currentDf = i
        pixelMove_nose_M1 = currentDf["Movement_mouse_1_nose"].mean()
        pixelMove_nose_M2 = currentDf["Movement_mouse_2_nose"].mean()
        mmMove_nose_M1 = pixelMove_nose_M1
        mmMove_nose_M2 = pixelMove_nose_M2
        list_nose_movement_M1.append(mmMove_nose_M1)
        list_nose_movement_M2.append(mmMove_nose_M2)
        current_velocity_M1_cm_sec = (mmMove_nose_M1)
        current_velocity_M2_cm_sec = (mmMove_nose_M2)
        current_velocity_M1_cm_sec = round(current_velocity_M1_cm_sec, 2)
        current_velocity_M2_cm_sec = round(current_velocity_M2_cm_sec, 2)
        meanVelocity_M1 = statistics.mean(list_nose_movement_M1)
        meanVelocity_M2 = statistics.mean(list_nose_movement_M2)
        meanVelocity_M1 = round(meanVelocity_M1,2)
        meanVelocity_M2 = round(meanVelocity_M2, 2)
        meanVelocity_M1_list.append(meanVelocity_M1)
        meanVelocity_M2_list.append(meanVelocity_M2)
        total_Movement_M1 = sum(list_nose_movement_M1)
        total_Movement_M2 = sum(list_nose_movement_M2)
        total_Movement_M1 = round(total_Movement_M1, 2)
        total_Movement_M2 = round(total_Movement_M2, 2)
        total_Movement_M1_list.append(total_Movement_M1)
        total_Movement_M2_list.append(total_Movement_M2)

        #save images
        for index, row in currentDf.iterrows():
            img_size = (400, 600, 3)
            img = np.ones(img_size) * 255
            cv2.putText(img, str('Mean velocity CD1: '), (5, 20), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (255, 0, 0), 1)
            cv2.putText(img, str('Mean velocity C57BL6/J: '), (5, 40), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (255, 0, 0), 1)
            cv2.putText(img, str('Total movement CD1: '), (5, 60), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (0, 0, 255), 1)
            cv2.putText(img, str('Total movement C57BL6/J: '), (5, 80), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (0, 0, 255), 1)
            cv2.putText(img, str('Current velocity CD1: '), (5, 100), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (0, 100, 0), 1)
            cv2.putText(img, str('Current velocity C57BL6/J: '), (5, 120), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (0, 100, 0), 1)
            cv2.putText(img, str(meanVelocity_M1) + str(' cm/s'), (275, 20), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (255, 0, 0), 1)
            cv2.putText(img, str(meanVelocity_M2) + str(' cm/s'), (275, 40), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (255, 0, 0), 1)
            cv2.putText(img, str(total_Movement_M1) + str(' cm'), (275, 60), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (0, 0, 255), 1)
            cv2.putText(img, str(total_Movement_M2) + str(' cm'), (275, 80), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (0, 0, 255), 1)
            cv2.putText(img, str(current_velocity_M1_cm_sec) + str(' cm/s'), (275, 100), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (0, 100, 0), 1)
            cv2.putText(img, str(current_velocity_M2_cm_sec) + str(' cm/s'), (275, 120), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (0, 100, 0), 1)
            centroid_distance_px = (int(row["Centroid_distance"]))
            centroid_distance_mm = (centroid_distance_px ) / 10
            centroid_distance_mm = round(centroid_distance_mm, 2)
            centroid_distance_mm_list.append(centroid_distance_mm)
            nose_2_nose_dist_px = (int(row["Nose_to_nose_distance"]))
            nose_2_nose_dist_mm = (nose_2_nose_dist_px ) / 10
            nose_2_nose_dist_mm = round(nose_2_nose_dist_mm, 2)
            nose_2_nose_dist_mm_list.append(nose_2_nose_dist_mm)
            cv2.putText(img, str('Centroid distance: '), (5, 140), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (153, 50, 204), 1)
            cv2.putText(img, str('Nose to nose distance: '), (5, 160), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (153, 50, 204), 1)
            cv2.putText(img, str(centroid_distance_mm) + str(' cm'), (275, 140), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (153, 50, 204), 1)
            cv2.putText(img, str(nose_2_nose_dist_mm) + str(' cm'), (275, 160), cv2.FONT_HERSHEY_TRIPLEX, 0.5, (153, 50, 204), 1)
            imageName = str(loop) + str('.png')
            imageSaveName = os.path.join(savePath, imageName)
            cv2.imwrite(imageSaveName, img)
            print('Processed ' + str(imageSaveName))
            loop += 1
            frameCounter += 1
    frames_processed_list.append(frameCounter)
    mean_centroid_distance_mm.append(statistics.mean(centroid_distance_mm_list))
    mean_nose_distance_mm.append(statistics.mean(nose_2_nose_dist_mm_list))
    meanVeloM1.append(statistics.mean(meanVelocity_M1_list))
    meanVeloM2.append(statistics.mean(meanVelocity_M2_list))
    totMoveM1.append(sum(total_Movement_M1_list))
    totMoveM2.append(sum(total_Movement_M2_list))


log_df['Video'] = VideoNo_list
log_df['Frames_processed'] = frames_processed_list
log_df['Mean_velocity_CD1'] = meanVeloM1
log_df['Mean_velocity_C57'] = meanVeloM2
log_df['Total_movement_CD1'] = total_Movement_M1
log_df['Total_movement_C57'] = total_Movement_M2
log_df['Mean_centroid_distance'] = mean_centroid_distance_mm
log_df['Mean_nose_to_nose_distance'] = mean_nose_distance_mm


book = load_workbook(log_fn)
writer = pd.ExcelWriter(log_fn, engine='openpyxl')
writer.book = book
log_df.to_excel(writer, sheet_name=sheetname, index=False)
writer.save()
writer.close()